#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
สรุปรีวิว -> report.xlsx  (รันหลัง scrape_reviews.py)
อ่าน data/reviews.csv + data/summary.csv แล้วสร้าง report.xlsx 4 ชีต:
  1) ภาพรวมต่อโรงแรม   - คะแนน, จำนวนรีวิวรวม, รีวิวใหม่รอบล่าสุด, รีวิวล่าสุด
  2) รีวิวใหม่รอบล่าสุด - เฉพาะรีวิวที่เพิ่งเข้ามาในรอบล่าสุด (เทียบรอบก่อน)
  3) รีวิวยอดแย่        - เรียงจากคะแนนต่ำสุด (ถ้ายังไม่มีดาว ใช้คะแนนลบจากคำเชิงลบ)
  4) รีวิวใหม่รายวัน    - นับรีวิวใหม่ต่อวันต่อโรงแรม

ใช้:  python analyze_reviews.py
"""
import csv, os, datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA = "data"
REVIEWS = os.path.join(DATA, "reviews.csv")
SUMMARY = os.path.join(DATA, "summary.csv")
OUT = "report.xlsx"

NEG_WORDS = ["สกปรก","แย่","เก่า","เสีย","ไม่สะอาด","เหม็น","หลอกลวง","มด","แมลง","ช้า",
             "ไม่แนะนำ","ผิดหวัง","ไม่ประทับใจ","ร้องเรียน","รั่ว","เชื้อรา","ขโมย","ค่าปรับ",
             "ไม่สุภาพ","ไม่ดี","ปรับปรุง","อันธพาล","ราคาแพง","น้ำไม่","ไม่เติม","ไม่ตรงปก","โกง"]


def read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def neg_score(text):
    t = text or ""
    return sum(t.count(w) for w in NEG_WORDS)


def rating_num(r):
    try:
        return float(r)
    except (TypeError, ValueError):
        return None


def main():
    reviews = read_csv(REVIEWS)
    summary = read_csv(SUMMARY)
    if not reviews:
        print("ไม่พบ data/reviews.csv"); return

    checked = sorted({r.get("CheckedAt", "") for r in reviews if r.get("CheckedAt")})
    latest = checked[-1] if checked else ""
    prev = checked[-2] if len(checked) > 1 else None
    new_rows = [r for r in reviews if r.get("CheckedAt") == latest]

    # ----- styling -----
    HDR = PatternFill("solid", fgColor="1F4D3A")
    HF = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D8D8D8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncol):
        for c in ws[1][:ncol]:
            c.fill = HDR; c.font = HF; c.border = border; c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    wb = openpyxl.Workbook()

    # ===== Sheet 1: ภาพรวมต่อโรงแรม =====
    ws = wb.active; ws.title = "ภาพรวมต่อโรงแรม"
    ws.append(["Code","โรงแรม","ทำเล","คะแนนเฉลี่ย","จำนวนรีวิวที่เก็บ","รีวิวใหม่รอบนี้","รีวิวล่าสุด"])
    by_hotel = defaultdict(list)
    for r in reviews:
        by_hotel[r["Code"]].append(r)
    smap = {s["Code"]: s for s in summary}
    for code, rs in sorted(by_hotel.items(), key=lambda kv: -float(smap.get(kv[0],{}).get("Rating") or 0)):
        s = smap.get(code, {})
        new_n = len([r for r in rs if r.get("CheckedAt") == latest])
        latest_txt = ""
        for r in rs:
            if r.get("CheckedAt") == latest and r.get("Text"):
                latest_txt = r["Text"][:200]; break
        ws.append([code, s.get("Hotel",""), s.get("Location",""), s.get("Rating",""),
                   len(rs), new_n, latest_txt])
    for row in ws.iter_rows(min_row=2):
        for c in row: c.border = border; c.alignment = wrap
    for col,w in zip("ABCDEFG",[8,30,14,11,14,13,60]):
        ws.column_dimensions[col].width = w
    style_header(ws, 7)

    # ===== Sheet 2: รีวิวใหม่รอบล่าสุด =====
    ws2 = wb.create_sheet("รีวิวใหม่รอบล่าสุด")
    note = "รอบนี้ (" + latest + ") มีรีวิวใหม่ " + str(len(new_rows)) + " รายการ"
    if prev: note += " | เทียบรอบก่อน " + prev
    ws2.append([note]); ws2.append([])
    ws2.append(["Code","โรงแรม","ผู้รีวิว","ดาว","ข้อความ"])
    hr = ws2.max_row
    for r in new_rows:
        ws2.append([r["Code"], r["Hotel"], r["Author"], r.get("Rating",""), r["Text"]])
    for row in ws2.iter_rows(min_row=hr):
        for c in row: c.border = border; c.alignment = wrap
    for col,w in zip("ABCDE",[8,26,18,7,90]):
        ws2.column_dimensions[col].width = w
    for c in ws2[hr]:
        c.fill = HDR; c.font = HF
    ws2["A1"].font = Font(bold=True, size=12, color="1F4D3A")

    # ===== Sheet 3: รีวิวยอดแย่ =====
    ws3 = wb.create_sheet("รีวิวยอดแย่")
    has_stars = any(rating_num(r.get("Rating")) is not None for r in reviews)
    ws3.append(["วิธีจัดอันดับ: " + ("ตามดาว 1-2 ก่อน" if has_stars else "ยังไม่มีดาวรายรีวิว → ใช้คะแนนลบจากคำเชิงลบในข้อความ")])
    ws3.append([])
    ws3.append(["Code","โรงแรม","ผู้รีวิว","ดาว","สัญญาณลบ","ข้อความ"])
    hr3 = ws3.max_row
    # เลือกรีวิวแย่สุด 3 อัน/โรงแรม
    for code, rs in by_hotel.items():
        scored = []
        for r in rs:
            rn = rating_num(r.get("Rating"))
            ns = neg_score(r.get("Text",""))
            # key: ดาวต่ำมาก่อน, ถ้าไม่มีดาวใช้คำลบมาก
            key = (rn if rn is not None else 99, -ns)
            if (rn is not None and rn <= 3) or (rn is None and ns >= 2):
                scored.append((key, ns, rn, r))
        scored.sort(key=lambda x: x[0])
        for key, ns, rn, r in scored[:3]:
            ws3.append([code, r["Hotel"], r["Author"], r.get("Rating",""), ns, r["Text"]])
    for row in ws3.iter_rows(min_row=hr3):
        for c in row: c.border = border; c.alignment = wrap
    for col,w in zip("ABCDEF",[8,24,18,7,9,80]):
        ws3.column_dimensions[col].width = w
    for c in ws3[hr3]:
        c.fill = PatternFill("solid", fgColor="B0573E"); c.font = HF
    ws3["A1"].font = Font(bold=True, italic=True, color="B0573E")

    # ===== Sheet 4: รีวิวใหม่รายวัน =====
    ws4 = wb.create_sheet("รีวิวใหม่รายวัน")
    ws4.append(["วันที่ตรวจ","Code","โรงแรม","จำนวนรีวิวใหม่"])
    daily = defaultdict(int); hotelname = {}
    for r in reviews:
        d = (r.get("CheckedAt","")[:10])
        daily[(d, r["Code"])] += 1
        hotelname[r["Code"]] = r["Hotel"]
    for (d, code), n in sorted(daily.items()):
        ws4.append([d, code, hotelname.get(code,""), n])
    for row in ws4.iter_rows(min_row=2):
        for c in row: c.border = border
    for col,w in zip("ABCD",[14,8,30,16]):
        ws4.column_dimensions[col].width = w
    style_header(ws4, 4)

    wb.save(OUT)
    print("saved " + OUT + " | new this run: " + str(len(new_rows)) + " | hotels: " + str(len(by_hotel)))


if __name__ == "__main__":
    main()
